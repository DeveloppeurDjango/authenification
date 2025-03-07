from django.shortcuts import render, redirect
from AppChefAgence.models import  OperateRetraite, ListRetraite
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import pandas as pd
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import locale
from django.core.paginator import Paginator
from django.core.files.storage import FileSystemStorage
import openpyxl
from datetime import datetime



def import_list_retraite_from_excel(file_path):
    # Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Parcourir les lignes du fichier Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ordre_enregistrement, titre, beneficiaire, nni, date_retraite, ville, age, telephone = row

        # Convertir la date au format approprié si nécessaire
        if isinstance(date_retraite, str):
            date_retraite = datetime.strptime(date_retraite, '%Y-%m-%d').date()

        # Créer un nouvel objet ListRetraite et sauvegarder dans la base de données
        list_retraite = ListRetraite(
            ordre=ordre_enregistrement,
            titre=titre,
            beneficiaire=beneficiaire,
            nni=nni,
            date_retraite=date_retraite,
            ville=ville,
            age=age,
            telephone=telephone
        )
        list_retraite.save()

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        
        # Appeler la fonction d'importation
        import_list_retraite_from_excel(file_path)
        
        return redirect('ajoutlistretraite')
    return render(request, 'upload_excel.html')



def Print(request, *args, **kwargs):
    # Spécifiez le chemin vers wkhtmltopdf
    
    operations = OperateRetraite.objects.all()
    
    total_montant = operations.aggregate(total_montant=Sum('montant'))['total_montant']

  
 
    
    template_path = 'Print.html'
   
    context = {'operate': operations, 'total_montant': total_montant}
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)
    
     # create a pdf
    pisa_status = pisa.CreatePDF(html,  dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response
  



def CreatedPDF(request, *args, **kwargs):
    # Spécifiez le chemin vers wkhtmltopdf
    
    operations = OperateRetraite.objects.all()
    
    total_montant = operations.aggregate(total_montant=Sum('montant'))['total_montant']

  
 
    
    template_path = 'Created_pdf.html'
   
    context = {'operate': operations, 'total_montant': total_montant}
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)
    
  

    # create a pdf
    pisa_status = pisa.CreatePDF(html,  dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response
   
from openpyxl.styles import Font, PatternFill

def ExportExcel(request):
    # Récupérer toutes les opérations
    operations = OperateRetraite.objects.all()

    # Calculer la somme des montants
    total_montant = operations.aggregate(total_montant=Sum('montant'))['total_montant']

    # Créer un DataFrame pandas à partir des données du modèle
    data = {
        'Ordre': [op.ordre for op in operations],
        'Titre': [op.Numero_titre for op in operations],
        'Bénéficiaire': [op.beneficiaire for op in operations],
        'NNI': [op.nni for op in operations],
        'Montant': [op.montant for op in operations],
    }

    df = pd.DataFrame(data)

    # Créer un DataFrame pour le montant total
    total_df = pd.DataFrame({'Ordre': ['Total'], 'Titre': [''], 'Bénéficiaire': [''], 'NNI': [''], 'Montant': [total_montant]})

    # Concaténer les deux DataFrames
    df = pd.concat([df, total_df], ignore_index=True)

    # Formater les colonnes Montant et NNI
    locale.setlocale(locale.LC_ALL, '')
    df['Montant'] = df['Montant'].apply(lambda x: locale.format_string("%d", x, grouping=True))

    # Utiliser un buffer en mémoire pour enregistrer le fichier Excel
    output = BytesIO()

    # Écrire le DataFrame dans le buffer en format Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Opérations')

        # Accéder au classeur et à la feuille de calcul
        workbook = writer.book
        worksheet = writer.sheets['Opérations']

        # Définir les styles de bordure
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        thick_border = Border(left=Side(style='medium'),
                              right=Side(style='medium'),
                              top=Side(style='medium'),
                              bottom=Side(style='medium'))

        # Appliquer les bordures aux cellules de titre
        for cell in worksheet[1]:
            cell.border = thick_border

        # Appliquer les bordures aux cellules de contenu
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Ajuster la taille des colonnes
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            # Si c'est la colonne "Bénéficiaire", ajustez la largeur à 30
            if column == 'C':
                adjusted_width = 30
            else:
                adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        # Centrer le contenu de toutes les cellules
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Appliquer le style gras à la cellule du montant total
        cell = worksheet.cell(row=len(df) + 1, column=5)
        cell.font = Font(bold=True)

        # Appliquer un fond marron-clair à la cellule du montant total
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.fill = fill

    # Réinitialiser la position du buffer à zéro
    output.seek(0)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=operations.xlsx'

    return response


